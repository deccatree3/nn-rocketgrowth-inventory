"""제품 마스터(WMS/쿠팡) 파일 파싱 및 DB upsert/replace.

파일 형식: 마스터-상품정보.xlsx 와 동일 구조
  WMS상품정보 시트 (헤더 row 2):
    WMS바코드 | 제품명 | 낱개수량 | 부모_WMS바코드 | 1카톤박스입수량 | 중량 | 소비기한일수 | 옵션ID | 부모_옵션ID
  쿠팡상품정보 시트 (헤더 row 1):
    등록상품ID | 옵션ID | SKU ID | 등록상품명 | 옵션명 | 상품등급 | 상품등록일 | 수동입고여부 | WMS바코드 | 쿠팡바코드 | WMS바코드-반품

단일 시트 파일도 자동 인식:
  - 첫 행에 'WMS바코드' 가 있으면 WMS 시트로 인식
  - 첫 행에 '옵션 ID' 또는 '옵션ID' 가 있으면 쿠팡 시트로 인식
"""
from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import delete, select

from lib.db import get_session
from lib.models import CoupangProduct, WmsProduct


def _to_int(v: Any) -> int | None:
    if v is None or v == "" or v == "-" or v == "#N/A":
        return None
    if isinstance(v, bool):
        return None
    try:
        return int(float(str(v).replace(",", "")))
    except (ValueError, TypeError):
        return None


def _to_str(v: Any) -> str | None:
    if v is None or v == "#N/A":
        return None
    s = str(v).strip()
    return s or None


def _to_date(v: Any):
    if v is None or v == "" or v == "#N/A":
        return None
    if isinstance(v, datetime):
        return v.date()
    if hasattr(v, "year"):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# 파싱
# ---------------------------------------------------------------------------
def parse_wms_sheet(ws) -> list[dict[str, Any]]:
    """openpyxl worksheet → WmsProduct 딕트 리스트.

    첫 컬럼이 '업체'이면 offset=1 (컬럼 한 칸 밀림), 아니면 offset=0.
    """
    # 헤더 행 + offset 감지
    header_row = 1
    offset = 0
    for r in range(1, min(5, ws.max_row + 1)):
        for c in range(1, min(4, ws.max_column + 1)):
            val = str(ws.cell(row=r, column=c).value or "")
            if "WMS" in val and "바코드" in val:
                header_row = r
                offset = c - 1  # '업체' 컬럼이 앞에 있으면 offset=1
                break
            if val.strip() == "바코드" or val.strip() == "WMS바코드":
                header_row = r
                offset = c - 1
                break
    start_row = header_row + 1
    o = offset  # shorthand

    records = []
    for r in range(start_row, ws.max_row + 1):
        company = _to_str(ws.cell(row=r, column=1).value) if o >= 1 else None
        barcode = _to_str(ws.cell(row=r, column=1 + o).value)
        if not barcode:
            continue
        rec = {
            "wms_barcode": barcode,
            "product_name": _to_str(ws.cell(row=r, column=2 + o).value),
            "unit_qty": _to_int(ws.cell(row=r, column=3 + o).value),
            "parent_wms_barcode": _to_str(ws.cell(row=r, column=4 + o).value),
            "box_qty": _to_int(ws.cell(row=r, column=5 + o).value),
            "weight_g": _to_int(ws.cell(row=r, column=6 + o).value),
            "shelf_life_days": _to_int(ws.cell(row=r, column=7 + o).value),
            "coupang_option_id": _to_int(ws.cell(row=r, column=8 + o).value),
            "parent_coupang_option_id": _to_int(ws.cell(row=r, column=9 + o).value),
        }
        if company:
            rec["company_name"] = company
        records.append(rec)
    return records


def parse_coupang_sheet(ws) -> list[dict[str, Any]]:
    """openpyxl worksheet → CoupangProduct 딕트 리스트.

    첫 컬럼이 '업체명'이면 offset=1, 아니면 offset=0.
    """
    header_row = 1
    offset = 0
    # 업체명 컬럼 감지
    for c in range(1, min(4, ws.max_column + 1)):
        val = str(ws.cell(row=1, column=c).value or "")
        if "업체" in val:
            offset = 1
            break
        if "등록상품" in val or "옵션" in val:
            offset = c - 1  # 등록상품 ID가 첫 컬럼이면 offset=0
            break
    start_row = header_row + 1
    o = offset

    records = []
    for r in range(start_row, ws.max_row + 1):
        company = _to_str(ws.cell(row=r, column=1).value) if o >= 1 else None
        option_id = _to_int(ws.cell(row=r, column=2 + o).value)
        if not option_id:
            continue
        flag = ws.cell(row=r, column=8 + o).value
        managed = flag in (1, "1") or (isinstance(flag, (int, float)) and int(flag) == 1)
        rec = {
            "coupang_option_id": option_id,
            "coupang_product_id": _to_int(ws.cell(row=r, column=1 + o).value),
            "sku_id": _to_int(ws.cell(row=r, column=3 + o).value),
            "product_name": _to_str(ws.cell(row=r, column=4 + o).value) or f"옵션 {option_id}",
            "option_name": _to_str(ws.cell(row=r, column=5 + o).value),
            "grade": _to_str(ws.cell(row=r, column=6 + o).value),
            "registered_at": _to_date(ws.cell(row=r, column=7 + o).value),
            "milkrun_managed": managed,
            "wms_barcode": _to_str(ws.cell(row=r, column=9 + o).value),
            "coupang_barcode": _to_str(ws.cell(row=r, column=10 + o).value),
            "wms_barcode_return": _to_str(ws.cell(row=r, column=11 + o).value),
            "active": True,
        }
        if company:
            rec["company_name"] = company
        records.append(rec)
    return records


def parse_master_file(file_bytes: bytes, filename: str) -> dict[str, list[dict]]:
    """마스터 파일을 파싱하여 {'wms': [...], 'coupang': [...]} 반환.

    양쪽 시트가 있으면 둘 다, 한쪽만 있으면 해당 것만 반환.
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    result: dict[str, list[dict]] = {"wms": [], "coupang": []}

    for sname in wb.sheetnames:
        ws = wb[sname]
        sname_upper = sname.upper()

        # 1) 시트 이름으로 우선 판별 (가장 신뢰도 높음)
        if "WMS" in sname_upper and "쿠팡" not in sname:
            result["wms"] = parse_wms_sheet(ws)
            continue
        if "쿠팡" in sname:
            result["coupang"] = parse_coupang_sheet(ws)
            continue

        # 2) 시트 이름으로 판별 안 되면 첫 행 내용으로 폴백
        first_row_vals = [str(ws.cell(row=1, column=c).value or "") for c in range(1, 5)]
        first_text = " ".join(first_row_vals)

        if first_text.startswith("WMS") or ("WMS바코드" == first_row_vals[0].strip()):
            result["wms"] = parse_wms_sheet(ws)
        elif "등록상품" in first_text or "옵션 ID" in first_text or "옵션ID" in first_text:
            result["coupang"] = parse_coupang_sheet(ws)

    wb.close()
    return result


# ---------------------------------------------------------------------------
# DB 적용
# ---------------------------------------------------------------------------
def upsert_wms_records(records: list[dict], replace_all: bool = False) -> dict[str, int]:
    """WMS 상품 upsert (또는 전체 교체). {added, updated, deleted} 반환."""
    from sqlalchemy.dialects.postgresql import insert

    stats = {"added": 0, "updated": 0, "deleted": 0}
    with get_session() as session:
        if replace_all:
            # 파일에 없는 기존 행 삭제
            file_barcodes = {r["wms_barcode"] for r in records}
            existing = session.execute(select(WmsProduct.wms_barcode)).scalars().all()
            to_delete = [bc for bc in existing if bc not in file_barcodes]
            if to_delete:
                session.execute(delete(WmsProduct).where(WmsProduct.wms_barcode.in_(to_delete)))
                stats["deleted"] = len(to_delete)

        existing_set = set(session.execute(select(WmsProduct.wms_barcode)).scalars().all())
        for rec in records:
            stmt = insert(WmsProduct).values(**rec)
            set_cols = {k: getattr(stmt.excluded, k) for k in rec if k != "wms_barcode"}
            set_cols["updated_at"] = datetime.now(timezone.utc)
            stmt = stmt.on_conflict_do_update(index_elements=["wms_barcode"], set_=set_cols)
            session.execute(stmt)
            if rec["wms_barcode"] in existing_set:
                stats["updated"] += 1
            else:
                stats["added"] += 1
        session.commit()
    return stats


def upsert_coupang_records(records: list[dict], replace_all: bool = False) -> dict[str, int]:
    """쿠팡 상품 upsert (또는 전체 교체). {added, updated, deleted} 반환."""
    from sqlalchemy.dialects.postgresql import insert

    stats = {"added": 0, "updated": 0, "deleted": 0}
    with get_session() as session:
        if replace_all:
            file_ids = {r["coupang_option_id"] for r in records}
            existing = session.execute(select(CoupangProduct.coupang_option_id)).scalars().all()
            to_delete = [oid for oid in existing if oid not in file_ids]
            if to_delete:
                session.execute(
                    delete(CoupangProduct).where(CoupangProduct.coupang_option_id.in_(to_delete))
                )
                stats["deleted"] = len(to_delete)

        existing_set = set(
            session.execute(select(CoupangProduct.coupang_option_id)).scalars().all()
        )
        for rec in records:
            stmt = insert(CoupangProduct).values(**rec)
            set_cols = {k: getattr(stmt.excluded, k) for k in rec if k != "coupang_option_id"}
            set_cols["updated_at"] = datetime.now(timezone.utc)
            stmt = stmt.on_conflict_do_update(
                index_elements=["coupang_option_id"], set_=set_cols
            )
            session.execute(stmt)
            if rec["coupang_option_id"] in existing_set:
                stats["updated"] += 1
            else:
                stats["added"] += 1
        session.commit()
    return stats

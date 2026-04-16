"""쿠팡 재고현황 파일(inventory_health_sku_info_*.xlsx) 파서.

파일 구조 (실측):
- 시트: 'Excel Data' (혹은 첫 시트)
- 1행: 대분류 헤더 (병합됨, 일부 셀에만 값)
- 2행: 소분류 헤더 (예: '최근 7일', '최근 30일', '1~30일' 등 서브 컬럼)
- 3행~: 데이터

컬럼 매핑 (1-indexed):
  1  No.
  2  등록상품 ID
  3  옵션 ID
  4  SKU ID
  5  등록상품명
  6  옵션명
  7  상품상태         ('NEW' 등)
  8  판매가능재고 (실시간 기준)
  9  입고중재고(실시간 기준)
 10  보관료구간
 11  최근 판매 (최근 7일)   -- 판매금액(원) 7일
 12  최근 판매 (최근 30일)  -- 판매금액(원) 30일
 13  최근 판매수량 (최근 7일)
 14  최근 판매수량 (최근 30일)
 15  추가입고 추천여부
 16  추가입고날짜 (입고마감일)
 17  권장재고 데이즈
 18  이번달 보관료(예상 기준)
 19  유통기간별 판매가능재고(박스기준) 1~30일
 20  31~45일
 21  46~60일
 22  61~120일
 23  121~180일
 24  181일+
 25  장기재고 최근 30일(박스기준)
 26  제조사명
 27  상품등록일
"""
from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .base import CoupangInventoryRow, CoupangSnapshot


def _to_int(v: Any) -> int:
    if v is None or v == "" or v == "-":
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).replace(",", "").strip()
    if not s or s == "-":
        return 0
    try:
        return int(float(s))
    except ValueError:
        return 0


def _to_int_opt(v: Any) -> int | None:
    if v is None or v == "" or v == "-":
        return None
    try:
        return int(float(str(v).replace(",", "")))
    except (ValueError, TypeError):
        return None


def _to_float_opt(v: Any) -> float | None:
    if v is None or v == "" or v == "-":
        return None
    try:
        return float(str(v).replace(",", ""))
    except (ValueError, TypeError):
        return None


def _to_str_opt(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


_DATE_IN_NAME = re.compile(r"inventory_health_sku_info_(\d{8})")


def _infer_snapshot_date(filename: str) -> date:
    m = _DATE_IN_NAME.search(filename)
    if m:
        try:
            return datetime.strptime(m.group(1), "%Y%m%d").date()
        except ValueError:
            pass
    return date.today()


def parse_coupang_inventory_file(path: str | Path) -> CoupangSnapshot:
    """쿠팡 재고현황 엑셀을 파싱하여 CoupangSnapshot 반환."""
    path = Path(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    rows: list[CoupangInventoryRow] = []
    # 1, 2행은 헤더, 3행부터 데이터
    for r in ws.iter_rows(min_row=3, values_only=True):
        if r is None or all(v is None or v == "" for v in r):
            continue
        # 옵션 ID 없으면 스킵
        option_id = _to_int_opt(r[2] if len(r) > 2 else None)
        if not option_id:
            continue

        row = CoupangInventoryRow(
            coupang_option_id=option_id,
            coupang_product_id=_to_int_opt(r[1] if len(r) > 1 else None),
            sku_id=_to_int_opt(r[3] if len(r) > 3 else None),
            product_name=_to_str_opt(r[4] if len(r) > 4 else None),
            option_name=_to_str_opt(r[5] if len(r) > 5 else None),
            orderable_stock=_to_int(r[7] if len(r) > 7 else None),
            inbound_stock=_to_int(r[8] if len(r) > 8 else None),
            sales_qty_7d=_to_int(r[12] if len(r) > 12 else None),
            sales_qty_30d=_to_int(r[13] if len(r) > 13 else None),
            recommendation=_to_str_opt(r[14] if len(r) > 14 else None),
            storage_fee_month=_to_float_opt(r[17] if len(r) > 17 else None),
            expiry_1_30=_to_int(r[18] if len(r) > 18 else None),
            expiry_31_45=_to_int(r[19] if len(r) > 19 else None),
            expiry_46_60=_to_int(r[20] if len(r) > 20 else None),
            expiry_61_120=_to_int(r[21] if len(r) > 21 else None),
            expiry_121_180=_to_int(r[22] if len(r) > 22 else None),
            expiry_181_plus=_to_int(r[23] if len(r) > 23 else None),
            raw={str(i): (str(v) if v is not None else None) for i, v in enumerate(r)},
        )
        rows.append(row)
    wb.close()

    return CoupangSnapshot(
        snapshot_date=_infer_snapshot_date(path.name),
        source_type="file",
        source_file=path.name,
        rows=rows,
    )

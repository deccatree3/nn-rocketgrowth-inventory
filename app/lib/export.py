"""쿠팡 로켓그로스 입고 업로드용 엑셀 생성.

두 가지 모드:
  1) fill_coupang_template(template_path, plan_items, ...)
     - 쿠팡 Wing에서 다운받은 빈 입고요청 양식(rows 5+)에 우리 plan 값(V, X~AA)을 채움
     - 쿠팡 바코드/상품ID 등은 원본 그대로 유지 → 업로드 호환 보장
  2) build_plain_xlsx(plan_items)
     - 디버그/미리보기 용도의 단순 xlsx (헤더 + 데이터만)

쿠팡 양식 컬럼 매핑 (1-indexed):
  A=No, B=등록상품명, C=옵션명, D=판매가, E=노출상품ID, F=등록상품ID, G=옵션ID
  H=판매방식
  I~U=판매이력/수수료 (쿠팡에서 자동 채움, 우리 건드리지 않음)
  V=입고 수량 입력 (필수)      ← 우리가 채움
  W=입고수량에 따른 예상매출    ← 파생, 안 건드림 (쿠팡이 계산)
  X=유통기간 입력(일수)        ← 필수 (우리 채움)
  Y=유통(소비)기한 (date)       ← 필수 (우리 채움)
  Z=제조일자 (date)            ← 필수 (우리 채움)
  AA=생산년도                   ← '선택입력' 문자열 또는 연도
  AB=상품바코드                 ← 원본 유지 (또는 WMS 바코드)
  AC=상품 사이즈                ← 'Small'/'Medium'/'Large'
  AD=취급주의여부               ← '해당아님' 기본
  AE-AM=기타 (원본 유지)
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

SHEET_NAME = "로켓그로스 입고"

# 1-indexed column numbers
COL_NO = 1
COL_PRODUCT_NAME = 2
COL_OPTION_NAME = 3
COL_PRICE = 4
COL_EXPOSED_PRODUCT_ID = 5
COL_REGISTERED_PRODUCT_ID = 6
COL_OPTION_ID = 7
COL_INBOUND_QTY = 22        # V
COL_EXPIRY_DAYS = 24        # X (유통기간 일수)
COL_EXPIRY_DATE = 25        # Y (유통기한 date)
COL_MANUFACTURE_DATE = 26   # Z (제조일자)
COL_PRODUCTION_YEAR = 27    # AA
COL_BARCODE = 28            # AB
COL_PRODUCT_SIZE = 29       # AC
COL_CAUTION = 30            # AD


@dataclass
class ExportItem:
    coupang_option_id: int
    inbound_qty: int
    shelf_life_days: int | None        # 유통기한 일수 (예: 730)
    expiry_date: date | None           # 유통(소비)기한
    manufacture_date: date | None      # 제조일자
    wms_barcode: str | None
    # 파생 가능 항목 (템플릿 모드에서는 원본 유지, plain 모드에서만 사용)
    product_name: str | None = None
    option_name: str | None = None


def fill_coupang_template(
    template_path: str | Path,
    items: list[ExportItem],
    default_size: str = "Small",
    default_caution: str = "해당아님",
    default_production_year: str = "선택입력",
    delete_non_target: bool = True,
) -> bytes:
    """쿠팡 다운로드 템플릿에 plan 값 주입하고 xlsx 바이트 반환.

    템플릿은 1~4행이 메타(제목/헤더/설명), 5행부터 옵션별 data row.
    옵션 ID로 매칭해서 해당 행의 V(입고수량), X~AA(유통기한/제조일) 을 채운다.

    delete_non_target=True (기본):
        입고 대상이 아닌 상품 행을 삭제하여 쿠팡 Wing 업로드용 깔끔한 파일 생성.
    """
    wb = load_workbook(template_path)
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        ws = wb.active

    # 옵션 ID → 해당 행 번호 인덱스
    id_to_row: dict[int, int] = {}
    for r in range(5, ws.max_row + 1):
        opt_id = ws.cell(row=r, column=COL_OPTION_ID).value
        try:
            key = int(str(opt_id)) if opt_id not in (None, "") else None
        except (ValueError, TypeError):
            key = None
        if key:
            id_to_row[key] = r

    # 입고 대상 옵션 ID 집합
    target_ids = {item.coupang_option_id for item in items}

    # 값 채우기 — 입고수량, 유통기한, 제조일자 3개만. 나머지는 원본 그대로 유지.
    missing: list[dict] = []
    filled = 0
    for item in items:
        r = id_to_row.get(item.coupang_option_id)
        if r is None:
            missing.append({
                "coupang_option_id": item.coupang_option_id,
                "product_name": item.product_name,
                "inbound_qty": item.inbound_qty,
            })
            continue
        ws.cell(row=r, column=COL_INBOUND_QTY).value = int(item.inbound_qty)
        if item.expiry_date is not None:
            ws.cell(row=r, column=COL_EXPIRY_DATE).value = item.expiry_date
        if item.manufacture_date is not None:
            ws.cell(row=r, column=COL_MANUFACTURE_DATE).value = item.manufacture_date
        filled += 1

    # 입고 대상이 아닌 행 삭제 (아래에서 위로 삭제하여 행 번호 shift 방지)
    if delete_non_target:
        rows_to_delete = sorted(
            [r for opt_id, r in id_to_row.items() if opt_id not in target_ids],
            reverse=True,
        )
        for r in rows_to_delete:
            ws.delete_rows(r, 1)

    buf = BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.getvalue(), missing


def build_plain_xlsx(items: list[ExportItem]) -> bytes:
    """템플릿 없이 단순 엑셀 생성 (디버그/미리보기용).

    헤더는 쿠팡 양식의 주요 컬럼만 포함한다.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    headers = [
        "No.",
        "등록상품명",
        "옵션명",
        "옵션 ID",
        "입고 수량",
        "유통기간(일)",
        "유통(소비)기한",
        "제조일자",
        "생산년도",
        "상품바코드",
        "상품 사이즈",
        "취급주의여부",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i).value = h

    for idx, item in enumerate(items, start=1):
        row = [
            idx,
            item.product_name or "",
            item.option_name or "",
            item.coupang_option_id,
            int(item.inbound_qty),
            str(item.shelf_life_days) if item.shelf_life_days else "",
            item.expiry_date,
            item.manufacture_date,
            "선택입력",
            item.wms_barcode or "",
            "Small",
            "해당아님",
        ]
        for i, v in enumerate(row, start=1):
            ws.cell(row=idx + 1, column=i).value = v

    # 열 너비
    widths = [6, 30, 20, 14, 10, 12, 14, 14, 12, 18, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.getvalue()


def default_expiry_dates(
    shelf_life_days: int | None,
    today: date | None = None,
) -> tuple[date | None, date | None]:
    """폴백용 — WMS 배치가 없을 때 사용하는 기본 유통기한/제조일 추정.

    제조일 = 오늘 - 7일 (임의)
    유통기한 = 제조일 + shelf_life_days
    """
    if not shelf_life_days:
        return None, None
    t = today or date.today()
    manufacture = t - timedelta(days=7)
    expiry = manufacture + timedelta(days=int(shelf_life_days) - 1)  # 제조일 포함이라 -1
    return expiry, manufacture


def dates_from_batch(
    batch_expiry: date | None,
    shelf_life_days: int | None,
) -> tuple[date | None, date | None]:
    """선택된 출고 배치의 유통일 기준으로 (유통기한, 제조일) 계산.

    제조일 = 유통기한 - shelf_life_days
    """
    if batch_expiry is None:
        return None, None
    if shelf_life_days and shelf_life_days > 0:
        manufacture = batch_expiry - timedelta(days=int(shelf_life_days) - 1)  # 제조일 포함 +1
    else:
        manufacture = None
    return batch_expiry, manufacture

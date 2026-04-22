"""밀크런 2차 결과물 3개 파일 생성.

1) 취합리스트 (서현_밀크런_취합리스트_*_FC.xlsx)
2) 팔레트적재리스트 (밀크런_물류부착문서2 (팔레트적재리스트)_FC_*.xlsx)
3) 재고이동건 (쿠팡 재고이동건_YYYY_MM월.xlsx) — 기존 파일에 새 시트 추가

모두 openpyxl 로 생성. 샘플 형식을 기준으로 컬럼 순서/위치 재현.
"""
from __future__ import annotations

import re
import zipfile
from dataclasses import dataclass, field
from datetime import date, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape as _xml_escape

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from lib.coupang_result import LabelInfo
from lib.pallet_assign import PalletAssignment, PalletEntry


# ============================================================================
# 공통 데이터 구조
# ============================================================================
@dataclass
class SecondaryItem:
    """2차 결과물 생성을 위한 SKU 단위 데이터."""

    coupang_option_id: int
    sku_id: int | None              # 거래명세서 상품번호 (= sku_id)
    coupang_product_id: int | None
    product_name: str               # 쿠팡 등록상품명 (팔레트적재리스트용)
    option_name: str | None
    own_wms_barcode: str | None
    coupang_barcode: str | None     # 쿠팡 부착바코드 (S00...)
    parent_wms_barcode: str | None
    unit_qty: int                   # 1, 2, 3, 6 ...
    inbound_qty: int                # 확정 수량 (낱개)
    box_qty: int                    # 박스 입수량
    boxes: int                      # 박스 수
    weight_g: int                   # 단위 중량 (g)
    expiry_date: date | None        # 소비기한
    manufacture_date: date | None   # 제조일자
    shelf_life_days: int | None     # 유통기한일수
    wms_product_name: str | None = None  # WMS 제품명 (취합리스트용)


def attached_barcode_and_type(item: SecondaryItem) -> tuple[str, str]:
    """이 SKU 의 부착바코드와 종류 반환."""
    if item.coupang_barcode and item.coupang_barcode.startswith("S0"):
        return item.coupang_barcode, "쿠팡바코드"
    return (item.own_wms_barcode or ""), "88코드"


def calc_weight_kg(item: SecondaryItem, boxes_in_section: int) -> float:
    """중량 = (단위중량 × 수량 + 500g × 박스수) / 1000 (kg)."""
    qty = boxes_in_section * item.box_qty
    return (item.weight_g * qty + 500 * boxes_in_section) / 1000


# ============================================================================
# 1) 취합리스트 (서현_밀크런_취합리스트_*_FC.xlsx)
# ============================================================================
def build_consolidation_list(
    items: list[SecondaryItem],
    pallet_assignment: PalletAssignment,
    fc_name: str,
    work_date: date,
    company_short: str = "서현",
    milkrun_id: str | None = None,
) -> bytes:
    """취합리스트 엑셀 생성.

    컬럼 (바코드종류 제외, '박스량' → '박스수'):
        1: 동탄1 / 상품명
        2: 바코드(WMS)
        3: 바코드(부착)
        4: 확정수량
        5: 소비기한
        6: 제조일자
        7: 박스 입수량
        8: 박스수
        9: 파레트번호
       10: 파레트수    ← 같은 팔레트 행들의 셀 병합
       11: 중량kg

    짝수 팔레트(2,4,...) 행은 옅은 빨강(살구색) 배경.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "상품리스트"

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    yellow = PatternFill("solid", fgColor="FFFF00")
    green = PatternFill("solid", fgColor="92D050")
    gray = PatternFill("solid", fgColor="E8E8E8")
    even_pallet_fill = PatternFill("solid", fgColor="FBE2D5")  # 옅은 빨강

    # ===== 상단 요약 헤더 (R1) — 회색 =====
    summary_headers = ["NO", "요청ID", "FC", "SKU 수", "총 수량", "총 박스수", "총 팔레트수", "총 중량"]
    for i, h in enumerate(summary_headers, start=4):  # D~K
        c = ws.cell(row=1, column=i, value=h)
        c.font = bold
        c.alignment = center
        c.border = border
        c.fill = gray

    # 합계 데이터 (R2)
    sku_count = sum(1 for it in items if it.boxes > 0)
    total_qty = sum(it.inbound_qty for it in items if it.boxes > 0)
    total_boxes = sum(it.boxes for it in items)
    total_pallets = pallet_assignment.pallet_count
    total_weight_kg = sum(
        (it.weight_g * it.inbound_qty + 500 * it.boxes) / 1000 for it in items if it.boxes > 0
    )

    row2 = [1, milkrun_id or "", fc_name, sku_count, total_qty, total_boxes, total_pallets, round(total_weight_kg, 2)]
    for i, v in enumerate(row2, start=4):
        c = ws.cell(row=2, column=i, value=v)
        c.alignment = center
        c.border = border

    # Total 행 (R5)
    tc = ws.cell(row=5, column=6, value="total")
    tc.font = bold
    tc.alignment = center
    for i, v in enumerate([sku_count, total_qty, total_boxes, total_pallets, round(total_weight_kg, 2)], start=7):
        c = ws.cell(row=5, column=i, value=v)
        c.alignment = center
        c.border = border
        c.font = bold

    # ===== 좌측 블록 헤더 (R8) — 팔레트 적재용 =====
    bundle_header = [
        fc_name,             # 1
        "바코드(WMS)",       # 2
        "바코드(부착)",      # 3
        "확정수량",          # 4
        "소비기한",          # 5
        "제조일자",          # 6
        "박스\n입수량",      # 7
        "박스수",            # 8
        "파레트번호",        # 9
        "파레트수",          # 10
        "중량kg",            # 11
    ]
    for i, h in enumerate(bundle_header, start=1):
        c = ws.cell(row=8, column=i, value=h)
        c.font = bold
        c.alignment = center
        c.border = border
        # 첫 3개는 초록, 나머지는 노랑 (샘플 패턴)
        c.fill = green if i <= 3 else yellow

    # ===== 우측 번들 정보 (N7 marker, R8 headers, R9~) =====
    # N7 (col 14, row 7) marker - 좌측 정렬
    marker = ws.cell(row=7, column=14, value="■ 번들작업표")
    marker.font = Font(bold=True, size=12)
    marker.alignment = Alignment(horizontal="left", vertical="center")

    bundle_right_headers = [
        ("바코드(WMS)", 14),
        ("바코드(부착)", 15),
        ("상품명", 16),
        ("수량", 17),
        ("소비기한", 18),
    ]
    for label, col in bundle_right_headers:
        c = ws.cell(row=8, column=col, value=label)
        c.font = bold
        c.alignment = center
        c.border = border
        c.fill = green if col <= 16 else yellow

    # ===== 데이터 (R9~) =====
    item_by_opt = {it.coupang_option_id: it for it in items}
    # (palette_no, entry, item) 순서대로 평탄화
    rows_to_write: list[tuple[int, PalletEntry, SecondaryItem]] = []
    for p_idx, pallet in enumerate(pallet_assignment.pallets, start=1):
        for entry in pallet:
            it = item_by_opt.get(entry.key)
            if it:
                rows_to_write.append((p_idx, entry, it))

    # 같은 팔레트별 시작 행을 추적해서 파레트수 셀 병합용
    pallet_row_ranges: dict[int, list[int]] = {}

    # 스키니퓨리티 선물세트 합포장 쇼핑백: 선물세트 바로 아래에 추가
    GIFT_SET_BC = "8809744301273"
    GIFT_BAG_BC = "8809744301525"
    GIFT_BAG_NAME = "스키니퓨리티 선물세트 쇼핑백 - 스키니퓨리티 선물세트(7T*4종) 박스에 합포장"

    r = 9
    for pallet_no, entry, it in rows_to_write:
        boxes_here = entry.boxes
        qty_here = boxes_here * it.box_qty
        weight = round((it.weight_g * qty_here + 500 * boxes_here) / 1000, 2)
        row_data = [
            it.wms_product_name or it.product_name,    # 1 WMS 제품명 우선
            it.own_wms_barcode or "",                 # 2
            attached_barcode_and_type(it)[0],         # 3
            qty_here,                                 # 4
            it.expiry_date.strftime("%Y-%m-%d") if it.expiry_date else "",   # 5
            it.manufacture_date.strftime("%Y-%m-%d") if it.manufacture_date else "",  # 6
            it.box_qty,                               # 7
            boxes_here,                               # 8
            pallet_no,                                # 9
            None,                                     # 10 — 병합 후 채움
            weight,                                   # 11
        ]
        for i, v in enumerate(row_data, start=1):
            c = ws.cell(row=r, column=i, value=v)
            # A열(1) 만 좌측 정렬, 나머지 가운데
            c.alignment = left_align if i == 1 else center
            c.border = border
            if pallet_no % 2 == 0:
                c.fill = even_pallet_fill

        pallet_row_ranges.setdefault(pallet_no, []).append(r)
        r += 1

        # 선물세트면 쇼핑백 합포장 행 1개 추가 (합계 / 팔레트 미반영)
        if str(it.own_wms_barcode or "") == GIFT_SET_BC:
            bag_row = [
                GIFT_BAG_NAME,  # 1 상품명
                GIFT_BAG_BC,    # 2 바코드(WMS)
                GIFT_BAG_BC,    # 3 바코드(부착)
                qty_here,       # 4 수량 = 선물세트 수량
                "", "", "", "", "", "", "",
            ]
            for i, v in enumerate(bag_row, start=1):
                c = ws.cell(row=r, column=i, value=v)
                c.alignment = left_align if i == 1 else center
                c.border = border
                if pallet_no % 2 == 0:
                    c.fill = even_pallet_fill
            r += 1

    # 파레트수 컬럼 (10) 병합 + 값 = 1
    for pallet_no, rows in pallet_row_ranges.items():
        if not rows:
            continue
        first, last = rows[0], rows[-1]
        # 첫 셀에 1
        ws.cell(row=first, column=10, value=1).alignment = center
        ws.cell(row=first, column=10).border = border
        if pallet_no % 2 == 0:
            ws.cell(row=first, column=10).fill = even_pallet_fill
        # 여러 행이면 병합
        if last > first:
            ws.merge_cells(start_row=first, end_row=last, start_column=10, end_column=10)

    # ===== 우측 번들 데이터 (R9~ col 14-18) =====
    bundle_items = [it for it in items if it.unit_qty and it.unit_qty >= 2 and it.boxes > 0]
    bundle_items.sort(key=lambda x: x.product_name or "")
    rb = 9
    for it in bundle_items:
        attached, _ = attached_barcode_and_type(it)
        right_data = [
            (it.own_wms_barcode or "", 14),
            (attached, 15),
            (it.wms_product_name or it.product_name, 16),
            (it.inbound_qty, 17),
            (it.expiry_date.strftime("%Y-%m-%d") if it.expiry_date else "", 18),
        ]
        for v, col in right_data:
            c = ws.cell(row=rb, column=col, value=v)
            # P열(16, 상품명) 좌측, 나머지 가운데
            c.alignment = left_align if col == 16 else center
            c.border = border
        rb += 1

    # ===== 컬럼 너비 (사용자 지정) =====
    # A, P = 70 / B, C, N, O = 16 / D~M, Q, R = 10
    width_map = {
        "A": 70, "B": 16, "C": 16,
        "D": 10.7, "E": 10.7, "F": 10.7, "G": 10.7, "H": 10.7,
        "I": 10.7, "J": 10.7, "K": 10.7, "L": 10.7, "M": 10.7,
        "N": 16, "O": 16, "P": 70, "Q": 10.7, "R": 10.7,
    }
    for col_letter, w in width_map.items():
        ws.column_dimensions[col_letter].width = w

    buf = BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.getvalue()


# ============================================================================
# 2) 팔레트적재리스트
# ============================================================================
def build_pallet_loading_list(
    items: list[SecondaryItem],
    pallet_assignment: PalletAssignment,
    fc_name: str,
    arrival_date: date,
    milkrun_request_id: str | None = None,
    company_full: str = "㈜서현커머스",
    pallet_size: int = 19,
) -> bytes:
    """팔레트적재리스트 엑셀 — 레퍼런스 서식 100% 재현.

    레퍼런스: sample/4. 2차결과물/밀크런_물류부착문서2 (팔레트적재리스트)_동탄1_20260414.xlsx
    - 전체 폰트 14pt (제목만 22pt bold)
    - 헤더 NOT bold
    - 한 섹션 = 27행 (라벨~빈 데이터 슬롯 포함)
    - 병합: A:F + G:M (라벨/날짜), A:M (업체/요청), C:K (상품명 헤더)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet3"

    thin = Side(border_style="thin", color="000000")
    no_side = Side(border_style=None)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)        # A,B,L,M,N용
    border_c = Border(left=thin, right=no_side, top=thin, bottom=thin)   # C용 (오른쪽 없음)
    border_hz = Border(left=no_side, right=no_side, top=thin, bottom=thin)  # D~K용 (가로만)
    font14 = Font(size=14)
    font14b = Font(bold=True, size=14)
    font22b = Font(bold=True, size=22)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_a = Alignment(horizontal="left", vertical="center", wrap_text=False)

    # ===== 컬럼 너비 (레퍼런스 동일) =====
    ws.column_dimensions["A"].width = 7.5
    ws.column_dimensions["B"].width = 18.375
    ws.column_dimensions["C"].width = 9.5
    # D~K: default (openpyxl default ~8.43)
    ws.column_dimensions["L"].width = 11.25
    # M: default
    ws.column_dimensions["N"].width = 19.5

    # ===== 확대/축소 비율 70% =====
    ws.sheet_view.zoomScale = 70

    # ===== R1: 제목 (A1:M2 병합, 22pt bold center) =====
    ws.merge_cells("A1:M2")
    title_cell = ws.cell(row=1, column=1, value="쿠팡 파렛트 적재리스트(필수작성)")
    title_cell.font = font22b
    title_cell.alignment = center
    ws.row_dimensions[1].height = 16.5
    ws.row_dimensions[2].height = 16.5

    item_by_opt = {it.coupang_option_id: it for it in items}
    total_pallets = pallet_assignment.pallet_count

    SECTION_HEIGHT = 27
    SECTION_START = 4
    MAX_DATA_ROWS = 17  # R(start+8+1) ~ R(start+8+17) = 빈 슬롯 포함

    for p_idx, pallet in enumerate(pallet_assignment.pallets, start=1):
        s = SECTION_START + (p_idx - 1) * SECTION_HEIGHT  # 섹션 시작 행
        boxes_in_pallet = sum(e.boxes for e in pallet)

        # --- 행 높이 (레퍼런스 동일) ---
        ws.row_dimensions[s - 1].height = 11.25 if s > 4 else 11.25  # gap
        ws.row_dimensions[s].height = 20.25      # 라벨
        ws.row_dimensions[s + 1].height = 11.25  # gap
        ws.row_dimensions[s + 2].height = 20.25  # 도착일
        ws.row_dimensions[s + 3].height = 11.25  # gap
        ws.row_dimensions[s + 4].height = 20.25  # 업체명
        ws.row_dimensions[s + 5].height = 11.25  # gap
        ws.row_dimensions[s + 6].height = 39.95  # 요청ID
        ws.row_dimensions[s + 7].height = 11.25  # gap
        ws.row_dimensions[s + 8].height = 44.25  # 헤더

        # --- 라벨 + 박스수량 (병합 A:F, G:M) ---
        ws.merge_cells(start_row=s, end_row=s, start_column=1, end_column=6)
        ws.merge_cells(start_row=s, end_row=s, start_column=7, end_column=13)
        ws.cell(row=s, column=1, value=f"{total_pallets}-{p_idx}").font = font14b
        ws.cell(row=s, column=1).alignment = left_a
        ws.cell(
            row=s, column=7,
            value=f" /      박스수량.   (      {boxes_in_pallet}   BOX)",
        ).font = font14b
        ws.cell(row=s, column=7).alignment = left_a

        # --- 도착일 + 납품센터 (병합) ---
        ws.merge_cells(start_row=s + 2, end_row=s + 2, start_column=1, end_column=6)
        ws.merge_cells(start_row=s + 2, end_row=s + 2, start_column=7, end_column=13)
        ws.cell(
            row=s + 2, column=1,
            value=f"물류센터 도착 예정일자.  ( {arrival_date.month:02d}월 {arrival_date.day:02d}일 )",
        ).font = font14b
        ws.cell(row=s + 2, column=1).alignment = left_a
        ws.cell(row=s + 2, column=7, value=f"납품센터명.  ( {fc_name} 센터 )").font = font14b
        ws.cell(row=s + 2, column=7).alignment = left_a

        # --- 업체명 (A:M 병합) ---
        ws.merge_cells(start_row=s + 4, end_row=s + 4, start_column=1, end_column=13)
        ws.cell(row=s + 4, column=1, value=f"업체명.  ( {company_full} )").font = font14b
        ws.cell(row=s + 4, column=1).alignment = left_a

        # --- 요청ID (A:M 병합) ---
        ws.merge_cells(start_row=s + 6, end_row=s + 6, start_column=1, end_column=13)
        if milkrun_request_id:
            ws.cell(row=s + 6, column=1, value=f"요청ID ({milkrun_request_id})").font = font14b
        ws.cell(row=s + 6, column=1).alignment = left_a

        # --- 헤더 (NOT bold, 14pt, center, border) ---
        hr = s + 8
        # C:K 병합 (상품명 헤더)
        ws.merge_cells(start_row=hr, end_row=hr, start_column=3, end_column=11)
        hdr_data = [
            (1, "NO"),
            (2, "거래명세서의\n상품번호"),
            (3, "물류 입고용 상품명 + 옵션명"),
            (12, "BOX수량"),
            (13, "수량"),
            (14, "소비기한/\n제조일자"),
        ]
        for col, label in hdr_data:
            c = ws.cell(row=hr, column=col, value=label)
            c.font = font14  # NOT bold (레퍼런스 동일)
            c.alignment = center
            c.border = border_c if col == 3 else border
        # 병합 안 된 헤더 셀(D~K): 가로만
        for col in range(4, 12):
            ws.cell(row=hr, column=col).border = border_hz

        # --- 데이터 행 (14pt, not bold, 행높이 34.5) ---
        for slot_idx in range(MAX_DATA_ROWS):
            dr = hr + 1 + slot_idx
            ws.row_dimensions[dr].height = 34.5

            if slot_idx < len(pallet):
                entry = pallet[slot_idx]
                it = item_by_opt.get(entry.key)
                if it:
                    qty = entry.boxes * it.box_qty
                    expiry_str = (
                        f"{it.expiry_date.strftime('%Y-%m-%d')}\n{it.manufacture_date.strftime('%Y-%m-%d')}"
                        if it.expiry_date and it.manufacture_date
                        else ""
                    )
                    name = ", ".join(filter(None, [it.product_name, it.option_name])).strip()

                    ws.cell(row=dr, column=1, value=slot_idx + 1).font = font14
                    ws.cell(row=dr, column=1).alignment = center
                    ws.cell(row=dr, column=2, value=str(it.sku_id) if it.sku_id else "").font = font14
                    ws.cell(row=dr, column=2).alignment = center
                    ws.cell(row=dr, column=3, value=name).font = font14
                    ws.cell(row=dr, column=3).alignment = left_a
                    ws.cell(row=dr, column=12, value=entry.boxes).font = font14
                    ws.cell(row=dr, column=12).alignment = center
                    ws.cell(row=dr, column=13, value=qty).font = font14
                    ws.cell(row=dr, column=13).alignment = center
                    ws.cell(row=dr, column=14, value=expiry_str).font = font14
                    ws.cell(row=dr, column=14).alignment = center

            # 테두리: A,B,L,M,N = 전체 / C = 오른쪽 없음 / D~K = 가로만
            for col in [1, 2, 12, 13, 14]:
                ws.cell(row=dr, column=col).border = border
            ws.cell(row=dr, column=3).border = border_c
            for col in range(4, 12):
                ws.cell(row=dr, column=col).border = border_hz

    buf = BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.getvalue()


# ============================================================================
# 3) 재고이동건 (월별 누적 파일에 시트 추가)
# ============================================================================
def update_inventory_movement(
    existing_xlsx_bytes: bytes,
    items: list[SecondaryItem],
    work_date: date,
    fc_name: str,
    company_short: str = "서현",
) -> bytes:
    """재고이동건: 숨겨진 'form' 시트를 복사 → 시트명 변경 → 입고수량 채움 → 불필요 행 삭제.

    form 시트 구조 (수식 포함):
        A: 바코드     B: 상품명     C: 출고수량(=SUMIFS 수식)
        D: 입고수량    E: #(unit_qty)  F: #(=D*E)  G: #(부모상품명)
    - 단품: C = SUMIFS($F:$F,$G:$G,B) → 자식 번들의 F합을 자동 집계
    - 번들: D = 우리가 채울 값, F = D*E (수식)

    절차:
      1. 'form' 시트 복사
      2. 시트명: MMDD(서현, 밀크런, FC)
      3. 번들 상품의 D열에 입고수량 채움 (바코드 매칭)
      4. 관련 없는 행 삭제 (입고도 출고도 없을 행)
      5. A~G 필터 적용
    """
    wb = openpyxl.load_workbook(BytesIO(existing_xlsx_bytes))
    sheet_name = f"{work_date.month:02d}{work_date.day:02d}({company_short}, 밀크런, {fc_name})"

    # 동일명 시트 제거
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    # form 시트 복사
    if "form" not in wb.sheetnames:
        raise ValueError("재고이동건 파일에 'form' 시트가 없습니다. form 시트를 만들어주세요.")
    form_ws = wb["form"]
    ws = wb.copy_worksheet(form_ws)
    ws.title = sheet_name
    ws.sheet_state = "visible"

    # 시트 순서: 맨 앞으로 이동 + 활성 시트로 설정
    wb.move_sheet(ws, offset=-(len(wb.sheetnames) - 1))
    wb.active = wb.sheetnames.index(ws.title)
    for other_ws in wb.worksheets:
        for sv in other_ws.views.sheetView:
            sv.tabSelected = other_ws.title == ws.title

    # 번들 입고수량 매핑
    bundle_qty_by_bc: dict[str, int] = {}
    for it in items:
        if it.boxes <= 0:
            continue
        if it.unit_qty and it.unit_qty >= 2 and it.own_wms_barcode:
            bundle_qty_by_bc[str(it.own_wms_barcode)] = it.inbound_qty

    # 관련 바코드 셋 — 필터 결과와 동일하게 행 숨기기용
    # C>0 이 될 바코드 = 번들의 부모 바코드 (SUMIFS가 >0 반환)
    # D에 값이 있는 바코드 = 번들 바코드
    relevant_barcodes: set[str] = set(bundle_qty_by_bc.keys())  # 번들 자신
    for it in items:
        if it.boxes <= 0:
            continue
        if it.unit_qty and it.unit_qty >= 2 and it.parent_wms_barcode:
            relevant_barcodes.add(str(it.parent_wms_barcode))  # 번들의 부모

    # D열에 번들 입고수량 채움 + 비관련 행 숨기기 (필터 결과 재현)
    # C열은 수식 전용 — 건드리지 않음
    for r in range(2, ws.max_row + 1):
        barcode_val = ws.cell(row=r, column=1).value
        if barcode_val is None:
            ws.row_dimensions[r].hidden = True
            continue
        bc_str = str(barcode_val).strip()
        if bc_str in bundle_qty_by_bc:
            ws.cell(row=r, column=4, value=bundle_qty_by_bc[bc_str])
        # 관련 없는 행 숨김 (필터 해제하면 보임)
        if bc_str not in relevant_barcodes:
            ws.row_dimensions[r].hidden = True

    # A~G 자동 필터 + 조건
    from openpyxl.worksheet.filters import FilterColumn, CustomFilter, CustomFilters

    last_row = ws.max_row
    ws.auto_filter.ref = f"A1:G{last_row}"

    # C열(colId=2): "#"과 1이상만 노출
    ws.auto_filter.filterColumn.append(
        FilterColumn(
            colId=2,
            customFilters=CustomFilters(
                _and=True,
                customFilter=[
                    CustomFilter(operator="notEqual", val="0"),
                    CustomFilter(operator="notEqual", val="-"),
                ],
            ),
        )
    )
    # D열(colId=3): 빈 값 제외
    ws.auto_filter.filterColumn.append(
        FilterColumn(
            colId=3,
            customFilters=CustomFilters(
                customFilter=[CustomFilter(operator="notEqual", val="")]
            ),
        )
    )

    buf = BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.getvalue()


# ============================================================================
# 4) 로켓그로스(서현커머스) 발주서양식
# ============================================================================
# 발주서양식은 이지어드민 재고차감용 가상 발주 — 실제 배송 안 함. 배송지 정보는
# FC 와 무관하게 고정값 사용 (수령인 이름 템플릿만 FC 별로 다름).
RECEIVER_NAME_FMT = "[밀크런]로켓그로스_{fc}"
RECEIVER_PHONE = "010-1234-1234"
RECEIVER_ADDRESS = "충남 아산시 염치읍 서원리 72-16 2층 다원로지스틱스 아산센터"


# ---------------------------------------------------------------------------
# openpyxl 3.x 는 문자열을 inlineStr 로 저장. 이지어드민 파서는 sharedStrings
# 만 인식하므로 생성 xlsx 를 후처리해서 inlineStr → sharedStrings 로 변환.
# ---------------------------------------------------------------------------
_INLINE_STR_RE = re.compile(
    r'<c\s+([^>]*?)t="inlineStr"([^>]*)>\s*<is>\s*<t(?:\s[^>]*)?>(.*?)</t>\s*</is>\s*</c>',
    re.DOTALL,
)


def _convert_inline_to_shared_strings(xlsx_bytes: bytes) -> bytes:
    """openpyxl 이 inlineStr 로 저장한 xlsx 를 sharedStrings 포맷으로 변환."""
    with zipfile.ZipFile(BytesIO(xlsx_bytes)) as zin:
        files = {name: zin.read(name) for name in zin.namelist()}

    # sharedStrings 가 이미 있으면 그대로 반환 (다른 경로 xlsx 호환)
    if "xl/sharedStrings.xml" in files and b"inlineStr" not in files.get(
        "xl/worksheets/sheet1.xml", b""
    ):
        return xlsx_bytes

    # 워크시트 파일별로 inlineStr 을 sharedStrings 로 치환
    shared: dict[str, int] = {}

    def _register(text: str) -> int:
        if text not in shared:
            shared[text] = len(shared)
        return shared[text]

    def _replace_cell(m: re.Match) -> str:
        pre_attrs = m.group(1) or ""
        post_attrs = m.group(2) or ""
        raw = m.group(3) or ""
        # <t> 내부 텍스트 unescape — 이미 XML 이스케이프된 상태
        text = (
            raw.replace("&lt;", "<")
            .replace("&gt;", ">")
            .replace("&quot;", '"')
            .replace("&apos;", "'")
            .replace("&amp;", "&")
        )
        idx = _register(text)
        attrs = (pre_attrs + post_attrs).strip()
        return f'<c {attrs} t="s"><v>{idx}</v></c>' if attrs else f'<c t="s"><v>{idx}</v></c>'

    for name in list(files.keys()):
        if name.startswith("xl/worksheets/") and name.endswith(".xml"):
            text = files[name].decode("utf-8")
            new_text = _INLINE_STR_RE.sub(_replace_cell, text)
            files[name] = new_text.encode("utf-8")

    if not shared:
        return xlsx_bytes

    # sharedStrings.xml 생성 (성공 샘플 포맷에 맞춤 — xml:space 속성 불필요)
    items_xml = "".join(
        f"<si><t>{_xml_escape(s)}</t></si>" for s, _ in sorted(shared.items(), key=lambda x: x[1])
    )
    total = len(shared)
    files["xl/sharedStrings.xml"] = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        f'count="{total}" uniqueCount="{total}">{items_xml}</sst>'
    ).encode("utf-8")

    # [Content_Types].xml 에 sharedStrings override 추가
    ct_name = "[Content_Types].xml"
    ct = files[ct_name].decode("utf-8")
    if "sharedStrings.xml" not in ct:
        override = (
            '<Override PartName="/xl/sharedStrings.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
        )
        ct = ct.replace("</Types>", f"{override}</Types>")
        files[ct_name] = ct.encode("utf-8")

    # workbook.xml.rels 에 sharedStrings relationship 추가
    rels_name = "xl/_rels/workbook.xml.rels"
    rels = files[rels_name].decode("utf-8")
    if "sharedStrings.xml" not in rels:
        # 기존 Relationship Id 중 최대값 찾아 +1
        ids = re.findall(r'Id="rId(\d+)"', rels)
        next_id = max([int(x) for x in ids], default=0) + 1
        new_rel = (
            f'<Relationship Id="rId{next_id}" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" '
            'Target="sharedStrings.xml"/>'
        )
        rels = rels.replace("</Relationships>", f"{new_rel}</Relationships>")
        files[rels_name] = rels.encode("utf-8")

    # 재압축
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in files.items():
            zout.writestr(name, data)
    buf.seek(0)
    return buf.getvalue()


def order_form_sequence(
    items: list[SecondaryItem],
    pallet_assignment: PalletAssignment,
) -> list[SecondaryItem]:
    """취합리스트와 동일한 순서로 SKU 1회씩 배열 (팔레트 순회 기준 첫 등장 순서)."""
    item_by_opt = {it.coupang_option_id: it for it in items}
    seen: set[int] = set()
    ordered: list[SecondaryItem] = []
    for pallet in pallet_assignment.pallets:
        for entry in pallet:
            if entry.key in seen:
                continue
            it = item_by_opt.get(entry.key)
            if it and it.inbound_qty > 0:
                seen.add(entry.key)
                ordered.append(it)
    return ordered


def build_order_form(
    items: list[SecondaryItem],
    fc_name: str,
    order_number_base: str,
    pallet_assignment: PalletAssignment | None = None,
) -> bytes:
    """로켓그로스(서현커머스) 발주서양식 생성.

    컬럼: 순서, 주문번호, 상품명, 수량, 수령인, 연락처, 주소, 비고
    주문번호 = f"{요청ID}_{seq}"
    행 순서 = 취합리스트와 동일 (팔레트 순회 기준 SKU 첫 등장 순서)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    receiver = RECEIVER_NAME_FMT.format(fc=fc_name)
    phone = RECEIVER_PHONE
    address = RECEIVER_ADDRESS

    headers = ["순서", "주문번호", "상품명", "수량", "수령인", "연락처", "주소", "비고"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)

    ordered = (
        order_form_sequence(items, pallet_assignment)
        if pallet_assignment else [it for it in items if it.inbound_qty > 0]
    )

    for seq, it in enumerate(ordered, start=1):
        row = seq + 1
        ws.cell(row=row, column=1, value=seq)
        ws.cell(row=row, column=2, value=f"{order_number_base}_{seq}" if order_number_base else str(seq))
        ws.cell(row=row, column=3, value=it.product_name or "")
        ws.cell(row=row, column=4, value=int(it.inbound_qty))
        ws.cell(row=row, column=5, value=receiver)
        ws.cell(row=row, column=6, value=phone)
        ws.cell(row=row, column=7, value=address)

    buf = BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    # 이지어드민 파서 호환 — inlineStr 을 sharedStrings 로 변환
    return _convert_inline_to_shared_strings(buf.getvalue())


# ============================================================================
# 5) 재고차감 — 확장주문검색 파싱 / 검수 / 3차결과물 생성
# ============================================================================
@dataclass
class OrderSearchRow:
    order_no: str
    mgmt_no: str
    product_name: str
    barcode: str | None
    qty: int


def parse_order_search_file(content: bytes) -> list[OrderSearchRow]:
    """확장주문검색_*.xls 파싱.

    컬럼 인덱스: 6=관리번호, 7=주문번호, 8=판매상품명, 9=바코드, 10=상품수량
    """
    import xlrd
    wb = xlrd.open_workbook(file_contents=content)
    ws = wb.sheet_by_index(0)
    rows: list[OrderSearchRow] = []
    for r in range(1, ws.nrows):
        mgmt = str(ws.cell_value(r, 6)).strip() if ws.ncols > 6 else ""
        order_no = str(ws.cell_value(r, 7)).strip() if ws.ncols > 7 else ""
        pname = str(ws.cell_value(r, 8)).strip() if ws.ncols > 8 else ""
        bc = str(ws.cell_value(r, 9)).strip() if ws.ncols > 9 else ""
        try:
            qty = int(float(ws.cell_value(r, 10))) if ws.ncols > 10 and ws.cell_value(r, 10) != "" else 0
        except (ValueError, TypeError):
            qty = 0
        if not order_no and not mgmt:
            continue
        rows.append(OrderSearchRow(
            order_no=order_no, mgmt_no=mgmt, product_name=pname,
            barcode=bc or None, qty=qty,
        ))
    return rows


@dataclass
class StockDeductionCheck:
    status: str
    issues: list[str] = field(default_factory=list)
    matched_pairs: list[dict] = field(default_factory=list)


def validate_order_search(
    order_rows: list[OrderSearchRow],
    items: list[SecondaryItem],
    order_number_base: str,
    pallet_assignment: PalletAssignment | None = None,
    invoice_qty_by_sku: dict[str, int] | None = None,
) -> StockDeductionCheck:
    """확장주문검색 파일을 쿠팡 결과(거래명세서) 또는 발주확정 데이터와 비교 검수.

    우선순위:
      1) invoice_qty_by_sku (쿠팡 거래명세서 상품번호→확정수량) 제공 시 → 쿠팡 결과 기준
      2) 없으면 items 의 inbound_qty (발주확정 계획 기준)
    """
    issues: list[str] = []
    matched: list[dict] = []
    order_items = (
        order_form_sequence(items, pallet_assignment)
        if pallet_assignment else [it for it in items if it.inbound_qty > 0]
    )
    if len(order_rows) != len(order_items):
        issues.append(f"행수 불일치: 업로드 {len(order_rows)}건 / 기준 {len(order_items)}건")
    by_order_no = {r.order_no: r for r in order_rows}

    basis_label = "쿠팡결과" if invoice_qty_by_sku else "발주확정"

    for seq, it in enumerate(order_items, start=1):
        expected = f"{order_number_base}_{seq}" if order_number_base else str(seq)
        # 기준 수량 결정
        if invoice_qty_by_sku is not None:
            key = str(it.sku_id) if it.sku_id else None
            basis_qty = invoice_qty_by_sku.get(key) if key else None
            if basis_qty is None:
                issues.append(f"쿠팡결과 매칭 실패 (sku_id={key}): {it.product_name}")
                basis_qty = int(it.inbound_qty)
        else:
            basis_qty = int(it.inbound_qty)

        r = by_order_no.get(expected)
        if r is None:
            issues.append(f"주문번호 누락: {expected} ({it.product_name})")
            continue
        if int(r.qty) != int(basis_qty):
            issues.append(f"수량 불일치 {expected}: 업로드 {r.qty} / {basis_label} {basis_qty}")
        matched.append({
            "순서": seq, "주문번호": expected, "관리번호": r.mgmt_no,
            "상품명": it.product_name,
            f"{basis_label}수량": basis_qty, "업로드수량": r.qty,
        })
    return StockDeductionCheck(
        status="ok" if not issues else "fail",
        issues=issues, matched_pairs=matched,
    )


def _invoice_no(mgmt_no: str) -> str:
    """관리번호 → 송장번호: 관리번호 뒤에 '000000' 부착."""
    return f"{mgmt_no}000000"


def build_shipping_bulk_form(order_rows: list[OrderSearchRow]) -> bytes:
    """배송일괄처리양식 — 1컬럼 (송장번호)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value="송장번호")
    for i, r in enumerate(order_rows, start=1):
        ws.cell(row=i + 1, column=1, value=_invoice_no(r.mgmt_no))
    buf = BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.getvalue()


def build_invoice_upload_form(order_rows: list[OrderSearchRow]) -> bytes:
    """송장업로드양식 — 택배사/송장번호/관리번호 (A, D, E)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value="택배사")
    ws.cell(row=1, column=4, value="송장번호")
    ws.cell(row=1, column=5, value="관리번호")
    for i, r in enumerate(order_rows, start=1):
        ws.cell(row=i + 1, column=1, value="CJ대한통운")
        ws.cell(row=i + 1, column=4, value=_invoice_no(r.mgmt_no))
        ws.cell(row=i + 1, column=5, value=r.mgmt_no)
    buf = BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.getvalue()

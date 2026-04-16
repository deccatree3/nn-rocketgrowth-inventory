"""마스터-상품정보.xlsx 에서 wms_product + coupang_product 로 이관.

시트 구조:
  WMS상품정보 (row 2 헤더):
    WMS바코드 | 제품명 | 낱개수량 | 부모_WMS바코드 | 1카톤박스입수량 | 중량 | 소비기한일수 | 옵션ID | 부모_옵션ID
  쿠팡상품정보 (row 1 헤더):
    등록상품ID | 옵션ID | SKU ID | 등록상품명 | 옵션명 | 상품등급 | 상품등록일 | 수동입고여부 | WMS바코드 | 쿠팡바코드 | WMS바코드-반품

사용법:
  python scripts/import_master_from_catalog.py ../마스터-상품정보.xlsx
"""
from __future__ import annotations

import argparse
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from lib.db import get_session  # noqa: E402
from lib.models import CoupangProduct, WmsProduct  # noqa: E402


def _to_int(v):
    if v is None or v == "" or v == "-" or v == "#N/A":
        return None
    if isinstance(v, bool):
        return None
    try:
        return int(float(str(v).replace(",", "")))
    except (ValueError, TypeError):
        return None


def _to_str(v):
    if v is None or v == "#N/A":
        return None
    s = str(v).strip()
    return s or None


def _to_date(v):
    if v is None or v == "" or v == "#N/A":
        return None
    if isinstance(v, datetime):
        return v.date()
    if hasattr(v, "year") and hasattr(v, "month"):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def load_wms_products(wb) -> list[dict[str, Any]]:
    target = None
    for name in wb.sheetnames:
        if "WMS" in name:
            target = name
            break
    if target is None:
        raise RuntimeError("WMS상품정보 시트를 찾을 수 없음")
    ws = wb[target]

    # 헤더 행 찾기 (row 1이 빈 경우 row 2)
    header_row = 1 if ws.cell(row=1, column=1).value == "WMS바코드" else 2
    start_row = header_row + 1

    records = []
    for r in range(start_row, ws.max_row + 1):
        barcode = _to_str(ws.cell(row=r, column=1).value)
        if not barcode:
            continue
        records.append(
            {
                "wms_barcode": barcode,
                "product_name": _to_str(ws.cell(row=r, column=2).value),
                "unit_qty": _to_int(ws.cell(row=r, column=3).value),
                "parent_wms_barcode": _to_str(ws.cell(row=r, column=4).value),
                "box_qty": _to_int(ws.cell(row=r, column=5).value),
                "weight_g": _to_int(ws.cell(row=r, column=6).value),
                "shelf_life_days": _to_int(ws.cell(row=r, column=7).value),
                "coupang_option_id": _to_int(ws.cell(row=r, column=8).value),
                "parent_coupang_option_id": _to_int(ws.cell(row=r, column=9).value),
            }
        )
    return records


def load_coupang_products(wb) -> list[dict[str, Any]]:
    target = None
    for name in wb.sheetnames:
        if "쿠팡" in name:
            target = name
            break
    if target is None:
        raise RuntimeError("쿠팡상품정보 시트를 찾을 수 없음")
    ws = wb[target]

    header_row = 1
    start_row = header_row + 1

    records = []
    for r in range(start_row, ws.max_row + 1):
        option_id = _to_int(ws.cell(row=r, column=2).value)
        if not option_id:
            continue
        flag = ws.cell(row=r, column=8).value
        managed = flag in (1, "1") or (isinstance(flag, (int, float)) and int(flag) == 1)
        records.append(
            {
                "coupang_option_id": option_id,
                "coupang_product_id": _to_int(ws.cell(row=r, column=1).value),
                "sku_id": _to_int(ws.cell(row=r, column=3).value),
                "product_name": _to_str(ws.cell(row=r, column=4).value) or f"옵션 {option_id}",
                "option_name": _to_str(ws.cell(row=r, column=5).value),
                "grade": _to_str(ws.cell(row=r, column=6).value),
                "registered_at": _to_date(ws.cell(row=r, column=7).value),
                "milkrun_managed": managed,
                "wms_barcode": _to_str(ws.cell(row=r, column=9).value),
                "coupang_barcode": _to_str(ws.cell(row=r, column=10).value),
                "wms_barcode_return": _to_str(ws.cell(row=r, column=11).value),
                "active": True,
            }
        )
    return records


def upsert_wms(records: list[dict]) -> int:
    from sqlalchemy.dialects.postgresql import insert

    with get_session() as session:
        for rec in records:
            stmt = insert(WmsProduct).values(**rec)
            set_cols = {k: getattr(stmt.excluded, k) for k in rec if k != "wms_barcode"}
            set_cols["updated_at"] = datetime.now(timezone.utc)
            stmt = stmt.on_conflict_do_update(
                index_elements=["wms_barcode"],
                set_=set_cols,
            )
            session.execute(stmt)
        session.commit()
    return len(records)


def upsert_coupang(records: list[dict]) -> int:
    from sqlalchemy.dialects.postgresql import insert

    with get_session() as session:
        for rec in records:
            stmt = insert(CoupangProduct).values(**rec)
            set_cols = {
                k: getattr(stmt.excluded, k) for k in rec if k != "coupang_option_id"
            }
            set_cols["updated_at"] = datetime.now(timezone.utc)
            stmt = stmt.on_conflict_do_update(
                index_elements=["coupang_option_id"],
                set_=set_cols,
            )
            session.execute(stmt)
        session.commit()
    return len(records)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("catalog_path", type=Path)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if not args.catalog_path.exists():
        print(f"[ERR] 파일 없음: {args.catalog_path}")
        sys.exit(1)

    print(f"[1/3] 로드: {args.catalog_path.name}")
    wb = openpyxl.load_workbook(args.catalog_path, data_only=True)

    wms_records = load_wms_products(wb)
    cp_records = load_coupang_products(wb)
    managed = sum(1 for r in cp_records if r["milkrun_managed"])
    with_box = sum(1 for r in wms_records if r["box_qty"])
    print(f"      WMS상품정보: {len(wms_records)}건 (box_qty 보유: {with_box})")
    print(f"      쿠팡상품정보: {len(cp_records)}건 (수동입고=1: {managed})")

    if args.dry_run:
        print("[2/3] DRY-RUN — DB 변경 없음")
        for r in wms_records[:2]:
            print("  wms:", r)
        for r in cp_records[:2]:
            print("  cp:", r)
        return

    print("[2/3] wms_product upsert")
    n_wms = upsert_wms(wms_records)
    print(f"      {n_wms} rows")

    print("[3/3] coupang_product upsert")
    n_cp = upsert_coupang(cp_records)
    print(f"      {n_cp} rows")

    # 요약
    with get_session() as s:
        from sqlalchemy import func, select

        wms_total = s.execute(select(func.count()).select_from(WmsProduct)).scalar() or 0
        cp_total = s.execute(select(func.count()).select_from(CoupangProduct)).scalar() or 0
        cp_managed = (
            s.execute(
                select(func.count())
                .select_from(CoupangProduct)
                .where(CoupangProduct.milkrun_managed.is_(True))
            ).scalar()
            or 0
        )
    print(f"\n[완료] wms_product: {wms_total}, coupang_product: {cp_total} (managed: {cp_managed})")


if __name__ == "__main__":
    main()
